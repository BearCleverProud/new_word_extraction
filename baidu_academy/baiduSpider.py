# -*- coding: utf-8 -*-  


'''
Python 3.x
无忧代理IP Created on 2018年05月11日
描述：本DEMO演示了使用爬虫（动态）代理IP请求网页的过程，代码使用了多线程
逻辑：每隔5秒从API接口获取IP，对于每一个IP开启一个线程去抓取网页源码
@author: www.data5u.com
'''
import requests;
import time;
import urllib3;

from selenium import webdriver
import time, re
from openpyxl import load_workbook,Workbook
import configparser
import os
import pickle

ips = [];

def get_conf(conf_name):
    '''
    获取配置文件内容
    :param conf_name:配置名称
    :return:配置内容
    '''
    curpath = os.path.dirname(os.path.realpath(__file__))
    os.path.join(curpath, 'conf.ini')
    conf = configparser.ConfigParser()
    conf.read('conf.ini')
    conf_info = conf.get('myconf', conf_name)
    return conf_info

def get_web_info(driver, name, index=0):
    '''
    爬去页面信息，并实现下一页跳转。当无跳转按钮时结束。
    :param driver: 浏览器实体
    :return:爬取信息列表
    '''
    web_info = []
    driver.get(driver.current_url)
    print(name + ' page ' + str(index+1))
    # 获取初始页面，用于之后跳转
    base_page = driver.current_window_handle
    paper_list = driver.find_elements_by_class_name('c_font')
    done_nothing = True
    for i in paper_list:
        done_nothing = False
        paper_title = i.text
        i.find_element_by_tag_name('a').click()
        # 获得当前打开的第一个窗口句柄
        window_1 = driver.current_window_handle
        # 获得当前打开的所有窗口的句柄
        windows = driver.window_handles
        # 切换到当前最新的窗口
        for current_window in windows:
            if current_window != window_1:
                driver.switch_to.window(current_window)
        try:
            # journal_title = driver.find_element_by_class_name('journal_title').text
            key_word = driver.find_element_by_class_name('kw_main').text
            # author = driver.find_element_by_class_name('author_text').text
            abstract = driver.find_element_by_class_name('abstract_wr').text.split("\n")[1]
            # print(paper_title, journal_title, key_word, author)
            web_info.append({"title": paper_title, "keyword": key_word, "abstract": abstract})
            time.sleep(5)
            driver.close()
        except:
            if "wappass" in driver.current_url:
                print('IP被封', driver.current_url)
                pickle.dump(web_info, open("data/crawled_data" + name + str(index) + ".pkl", "wb"))
                return web_info, driver, False
            else:
                print("详情页错误", driver.current_url)
        driver.switch_to.window(base_page)
    # 查看是否有下一页
    
    try:
        next_page = driver.find_elements_by_id("page")[0]
        next_page.find_elements_by_tag_name('a')[-1].click()
        time.sleep(1)
    except:
        print('finish!!!')
        # driver.close()
        if not done_nothing:
            pickle.dump(web_info, open("data/crawled_data" + name + str(index) + ".pkl", "wb"))
        return web_info, driver, True
    if index % 10 == 0:
        pickle.dump(web_info, open("data/crawled_data" + name + str(index) + ".pkl", "wb"))
    new_info, driver, finished = get_web_info(driver, name, index+1)
    web_info += new_info
    # driver.close()
    return web_info, driver, finished

def get_name_list(file_loc):
    '''
    从excel中获取作者和作者对应的网址
    :param file_loc:excel文件地址，在配置文件中配置
    :return:作者和作者对应的网址
    '''
    workbook = load_workbook(file_loc)
    sheet1 = workbook.get_sheet_by_name('Sheet1')
    names = [i.value for i in sheet1['A']]
    sites = [i.value for i in sheet1['B']]
    name_site = [names, sites]
    # print(name_site)
    return name_site

def crawl(name, driver):
    print(name, '开始')
    # 访问出现空白页面时重新加载
    # while driver.find_elements_by_class_name('c_font') == []:
    #     print(name, '页面空白刷新')
    #     driver.get(url)
    info, driver, finished = get_web_info(driver, name)
    return driver, finished


# 爬数据的线程类
class CrawlThread():
    def __init__(self,proxyip):
        super(CrawlThread, self).__init__();
        self.proxyip=proxyip;
    def run(self, name, driver):
        # 开始计时
        # start = time.time();
        #消除关闭证书验证的警告
        urllib3.disable_warnings();
        #使用代理IP请求网址，注意第三个参数verify=False意思是跳过SSL验证（可以防止报SSL错误）
        return crawl(name, driver)
        # 结束计时
        # end = time.time();
        # 输出内容
        # print(threading.current_thread().getName() +  "使用代理IP, 耗时 " + str(end - start) + "毫秒 " + self.proxyip)

# 获取代理IP的线程类
class GetIpThread():
    def __init__(self,fetchSecond):
        super(GetIpThread, self).__init__();
        self.fetchSecond=fetchSecond;
    def run(self):
        global ips;
        for j in range(len(name_list[0])):
            init_driver = webdriver.Chrome()
            name = name_list[0][j]
            url = name_list[1][j]
            init_driver.get(url)
            finished = False
            while not finished:
                # 获取IP列表
                res = requests.get(apiUrl).content.decode()
                # 按照\n分割获取到的IP
                ips = res.split('\n');
                # 利用每一个IP
                for proxyip in ips:
                    if proxyip.strip():
                        # 开启一个线程
                        print("New IP: " + proxyip)
                        init_driver, finished = CrawlThread(proxyip).run(name, init_driver)
                # 休眠
                time.sleep(self.fetchSecond);
            init_driver.close()

if __name__ == '__main__':
    # 这里填写无忧代理IP提供的API订单号（请到用户中心获取）
    order = "ebb758478262d6dc0a2613f67b713998";
    # 获取IP的API接口
    apiUrl = "http://api.goubanjia.com/dynamic/get/ebb758478262d6dc0a2613f67b713998.html?sep=3";
    # 要抓取的目标网站地址
    # targetUrl = "http://pv.sohu.com/cityjson?ie=utf-8";
    # 获取IP时间间隔，建议为5秒
    fetchSecond = 5;
    # 开始自动获取IP
    file_loc = get_conf('file_loc')
    name_list = get_name_list(file_loc)

    GetIpThread(fetchSecond).run();



# file_loc = get_conf('file_loc')
# name_list = get_name_list(file_loc)
# # url = get_conf('url')
# # print(url)
# # wb = Workbook()
# # sheet = wb.active
# for j in range(len(name_list[0])):
#     init_driver = webdriver.Chrome()
#     name = name_list[0][j]
#     url = name_list[1][j]
#     print(name, '开始')
#     init_driver.get(url)
#     # 访问出现空白页面时重新加载
#     while init_driver.find_elements_by_class_name('c_font') == []:
#         print(name, '页面空白刷新')
#         init_driver.get(url)
#     get_web_info(init_driver, name)
    # sheet.title = "Sheet1"
    # for i in info:
    #     i += [name]
    #     sheet.append(i)
# wb.save(r'paper_list.xlsx')
